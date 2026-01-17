from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, current_app

from flask_login import login_required, current_user, login_user

from app.extensions import db, csrf

from app.models import (

    User, Location, Doctor, Service, AdditionalService, ServicePrice, AdditionalServicePrice,

    Clinic, Manager, PaymentMethod, Appointment, Organization, GlobalSetting, BonusPeriod,

    AppointmentHistory, AppointmentAdditionalService, AppointmentService, BonusValue, SystemMetrics,
    
    MedicalCertificate, Notification, NotificationStatus

)

from app.telegram_bot import telegram_bot

import psutil

from datetime import datetime, timedelta

from werkzeug.security import generate_password_hash, check_password_hash

import os



admin = Blueprint('admin', __name__)

import csv

import openpyxl

from werkzeug.utils import secure_filename

import io

import calendar

from datetime import date



admin = Blueprint('admin', __name__, template_folder='templates')



@admin.before_request
@login_required
def require_admin():
    # Allow org users specifically for their stats API, everyone else must be superadmin
    if current_user.role == 'org' and request.endpoint == 'admin.reports_organizations_details':
        return

    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))



@admin.route('/additional')

def additional():

    managers = Manager.query.all()

    payment_methods = PaymentMethod.query.all()

    centers = Location.query.filter_by(type='center').all()

    chat_setting = GlobalSetting.query.get('chat_image')

    chat_image = chat_setting.value if chat_setting else None
    
    stamp_setting = GlobalSetting.query.get('stamp_image')
    
    stamp_image = stamp_setting.value if stamp_setting else None

    return render_template('admin_additional.html', managers=managers, payment_methods=payment_methods, centers=centers, chat_image=chat_image, stamp_image=stamp_image)



@admin.route('/journal/clear', methods=['POST'])

@login_required

def clear_journal_data():

    center_id = request.form.get('center_id')

    month_str = request.form.get('month')

    password = request.form.get('password')

    

    if not center_id or not month_str:

        flash('Центр и Месяц обязательны', 'error')

        return redirect(url_for('admin.additional'))

        

    if not password or not check_password_hash(current_user.password_hash, password):

        flash('Неверный пароль', 'error')

        return redirect(url_for('admin.additional'))



    try:

        year, month = map(int, month_str.split('-'))

        start_date = date(year, month, 1)

        # Handle end of month

        _, last_day = calendar.monthrange(year, month)

        end_date = date(year, month, last_day)

        

        # Delete appointments

        num_deleted = Appointment.query.filter(

            Appointment.center_id == int(center_id),

            Appointment.date >= start_date,

            Appointment.date <= end_date

        ).delete(synchronize_session=False)

        

        db.session.commit()

        flash(f'Удалено записей: {num_deleted}', 'success')

        

    except Exception as e:

        db.session.rollback()

        flash(f'Ошибка: {str(e)}', 'error')



    return redirect(url_for('admin.additional'))



@admin.route('/journal/recalculate', methods=['POST'])

@login_required

def recalculate_journal_data():

    center_id = request.form.get('center_id')

    month_str = request.form.get('month')

    

    if not center_id or not month_str:

        flash('Центр и Месяц обязательны', 'error')

        return redirect(url_for('admin.additional'))

        

    try:

        year, month = map(int, month_str.split('-'))

        start_date = date(year, month, 1)

        _, last_day = calendar.monthrange(year, month)

        end_date = date(year, month, last_day)



        appointments = Appointment.query.filter(

            Appointment.center_id == int(center_id),

            Appointment.date >= start_date,

            Appointment.date <= end_date

        ).all()

        

        updated_count = 0

        

        for appt in appointments:

            if not appt.service: continue

            

            # Find service object by name to get price logic

            # Note: appt.service is string name. 

            service_obj = Service.query.filter(Service.name.ilike(appt.service)).first()

            

            price = 0.0

            if service_obj:

               price = service_obj.get_price(appt.date)

               

            # Additional services cost

            add_svc_cost = 0.0

            for add_svc in appt.additional_services:

                add_svc_cost += add_svc.get_price(appt.date)

                

            qty = appt.quantity if appt.quantity else 1

            add_qty = appt.additional_service_quantity if appt.additional_service_quantity else 1

            

            discount = appt.discount if appt.discount else 0.0

            

            # Recalculate Total

            # Formula: (ServicePrice * Qty) + (AddServicePriceTotal * AddQty) - Discount

            # Note: AddServicePriceTotal is sum of unit prices of all attached additional services

            

            # Wait, additional_services is a list. Do we multiply sum by add_qty OR 

            # does add_qty apply to each? 

            # In API create_appointment: 

            # add_services_cost = sum(add_svc.get_price(appt.date) for add_svc in appt.additional_services)

            # total_cost = (service_price * appt.quantity) + (add_services_cost * appt.additional_service_quantity)

            # This matches plan.

            

            new_cost = (price * qty) + (add_svc_cost * add_qty) - discount

            if new_cost < 0: new_cost = 0.0

            

            appt.cost = new_cost

            updated_count += 1

            

        db.session.commit()

        flash(f'Пересчитано записей: {updated_count}', 'success')



    except Exception as e:

        db.session.rollback()

        flash(f'Ошибка: {str(e)}', 'error')

        

    return redirect(url_for('admin.additional'))



@admin.route('/journal/import', methods=['POST'])

@login_required

def import_journal_data():

    center_id = request.form.get('center_id')

    if not center_id:

        flash('Не выбран центр', 'error')

        return redirect(url_for('admin.additional'))



    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.additional'))

    

    file = request.files['file']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.additional'))

        

    original_filename = file.filename

    if not (original_filename.lower().endswith('.csv') or original_filename.lower().endswith('.xlsx')):

        flash('Неподдерживаемый формат. Используйте CSV или XLSX.', 'error')

        return redirect(url_for('admin.additional'))

        

    filename = secure_filename(file.filename)



    try:

        rows = []

        if original_filename.lower().endswith('.csv'):

            content = file.stream.read()

            text = None

            try:

                text = content.decode("utf-8-sig")

            except UnicodeDecodeError:

                try:

                    text = content.decode("cp1251")

                except UnicodeDecodeError:

                    text = content.decode("utf-8", errors="ignore")

            

            stream = io.StringIO(text, newline=None)

            stream.seek(0)

            

            try:

                sample = stream.read(2048)

                stream.seek(0)

                if not sample: 

                    rows = []

                else:

                    dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")

                    csv_input = csv.reader(stream, dialect)

                    rows = list(csv_input)

            except csv.Error:

                stream.seek(0)

                csv_input = csv.reader(stream)

                rows = list(csv_input)

            except Exception:

                stream.seek(0)

                rows = list(csv.reader(stream))

            

        elif original_filename.lower().endswith('.xlsx'):

            wb = openpyxl.load_workbook(file, data_only=True)

            sheet = wb.active

            for r in sheet.iter_rows(values_only=True):

                rows.append(list(r))

        

        if not rows:

            flash(f'Файл пуст (Filename: {original_filename})', 'error')

            return redirect(url_for('admin.additional'))

            

        delete_old = request.form.get('delete_old') == 'on'



        # Header Detection

        headers = []

        if rows:

            headers = [str(h).strip().lower() if h else '' for h in rows[0]]

        

        header_row_idx = 0

        if 'дата' not in headers and len(rows) > 1:

             headers = [str(h).strip().lower() if h else '' for h in rows[1]]

             header_row_idx = 1

             

        col_map = {}

        for idx, h in enumerate(headers):

            if 'дата' in h and 'рождения' not in h: col_map['date'] = idx

            elif 'пациент' in h: col_map['patient'] = idx

            elif 'врач' in h: col_map['doctor'] = idx

            elif 'исследование' in h or 'услуга' in h: col_map['service'] = idx

            elif 'оплата' in h: col_map['payment'] = idx

            elif 'скидка' in h: col_map['discount'] = idx

            elif 'сумма' in h: col_map['cost'] = idx

            elif 'стоимость' in h and 'cost' not in col_map: col_map['cost'] = idx 

            elif 'договор' in h: col_map['contract'] = idx

            elif 'ребенок' in h: col_map['is_child'] = idx

            elif 'клиника' in h: col_map['clinic'] = idx

            elif 'кол-во' in h: col_map['quantity'] = idx

            elif 'доп' in h and 'услуги' in h: col_map['additional_services'] = idx

            elif 'комментарий' in h: col_map['comment'] = idx

            

        if 'date' not in col_map or 'patient' not in col_map:

            col_map = {

                'date': 0, 'contract': 3, 'patient': 4, 'is_child': 5, 'doctor': 6,

                'clinic': 8, 'service': 9, 'additional_services': 10, 'quantity': 11,

                'payment': 14, 'discount': 15, 'cost': 16, 'comment': 12

            }



        valid_rows_to_insert = []

        unique_dates_to_clean = set()

        warnings = []



        for i in range(header_row_idx + 1, len(rows)):

            row = rows[i]

            if not row or len(row) < 5: continue

            

            def get_val(key):

                idx = col_map.get(key)

                if idx is not None and idx < len(row):

                   return row[idx]

                return None



            date_val = get_val('date')

            if not date_val: continue

            

            appt_date = None

            if isinstance(date_val, datetime): appt_date = date_val.date()

            elif isinstance(date_val, date): appt_date = date_val

            elif isinstance(date_val, str): 

                val_str = date_val.strip()

                if not val_str: continue

                # Remove timestamps if present in string like "2025-12-01 00:00:00"

                # Replace comma with dot (e.g. 09,12.2025 -> 09.12.2025)

                val_str = val_str.replace(',', '.')

                

                for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y', '%d/%m/%Y']:

                    try:

                        appt_date = datetime.strptime(val_str, fmt).date()

                        break

                    except: pass

            

            if not appt_date: 

                warnings.append(f"Строка {i+1}: Неверный формат даты '{date_val}'.")

                continue

            

            patient = str(get_val('patient') or '').strip()

            if not patient: continue

            

            # --- STRICT LOGIC ---

            

            # 1. DOCTOR MATCHING

            doctor_name = str(get_val('doctor') or '').strip()

            

            # New Rule: Empty doctor -> "Без врача"

            if not doctor_name:

                doctor_name = "Без врача"

            

            doc_obj = None

            if doctor_name:

                doc_obj = Doctor.query.filter(Doctor.name.ilike(doctor_name)).first()

            

            if not doc_obj:

                # 1b. Fuzzy fallback: Clean spaces and try again

                import re

                

                # Normalize spaces (e.g. "Name  Surname" -> "Name Surname")

                clean_name = re.sub(r'\s+', ' ', doctor_name)

                if clean_name != doctor_name:

                    doc_obj = Doctor.query.filter(Doctor.name.ilike(clean_name)).first()

                

                # 1c. Super-fuzzy fallback: Replace spaces with wildcards

                if not doc_obj:

                    # "Ivanov Ivan" -> "Ivanov%Ivan"

                    wildcard_name = clean_name.replace(' ', '%')

                    doc_obj = Doctor.query.filter(Doctor.name.ilike(wildcard_name)).first()



            if not doc_obj:

                warnings.append(f"Строка {i+1}: Врач '{doctor_name}' не найден (даже с нечетким поиском).")

                continue # Skip row

                

            # 2. MANAGER AUTO-ASSIGN

            manager_id = None

            if doc_obj.manager:

                mgr = User.query.filter(User.username.ilike(doc_obj.manager)).first()

                if mgr: manager_id = mgr.id

            

            # 3. CLINIC MATCHING

            clinic_name = str(get_val('clinic') or '').strip()

            clinic_obj = None

            if clinic_name:

                 clinic_obj = Clinic.query.filter(Clinic.name.ilike(clinic_name)).first()

                 

            # Warning if clinic provided but not found? Or just ignore? 

            # Plan said "Match Clinic (skip if not found) or Warning".

            # Let's skip if clinic name was provided but not found, to be safe.

            # 4. SERVICE MATCHING

            service_name = str(get_val('service') or '').strip()

            service_obj = None

            

            if service_name:

                service_obj = Service.query.filter(Service.name.ilike(service_name)).first()

                if not service_obj:

                     warnings.append(f"Строка {i+1}: Услуга '{service_name}' не найдена (игнорируется, так как не найдена в каталоге).")

                     # If service name was present but invalid, should we treat it as "No Service" or "Skip"?

                     # Strict mode usually implies Skip. But let's see if we have Add Services.

                     # If user says "Load data if Add Srv not empty", maybe we treat invalid main service as None?

                     # Let's keep strictness: If valid name provided but not found -> Warning (and effectively None for logic below).

                     pass



            # 5. ADDITIONAL SERVICES MATCHING

            add_services_str = str(get_val('additional_services') or '').strip()

            add_service_objs = []

            if add_services_str:

                # Split by comma

                for raw_name in add_services_str.split(','):

                    name = raw_name.strip()

                    if not name: continue

                    

                    add_srv = AdditionalService.query.filter(AdditionalService.name.ilike(name)).first()

                    if not add_srv:

                         warnings.append(f"Строка {i+1}: Доп. услуга '{name}' не найдена.")

                         # Strict: If explicitly listed but not found, we skip the row to avoid data loss/corruption.

                         add_service_objs = [] # clear to trigger skip

                         service_obj = None # ensure skip

                         break

                    

                    add_service_objs.append(add_srv)



            # VALIDATION: Must have either Main Service OR Additional Services

            if not service_obj and not add_service_objs:

                 if not warnings: # Don't duplicate if we already warned about invalid service

                     warnings.append(f"Строка {i+1}: Не указана ни основная, ни доп. услуга.")

                 continue



            # COST CALCULATION

            # Parse quantity first for cost calc

            excel_qty = 1

            raw_qty = get_val('quantity')

            if raw_qty:

                try: excel_qty = int(float(str(raw_qty).replace(',', '.')))

                except: excel_qty = 1



            cost = 0.0

            if service_obj:

                cost += service_obj.get_price(appt_date) # Main service always quantity 1

            

            for ads in add_service_objs:

                cost += ads.get_price(appt_date) * excel_qty



            

            # --- END STRICT LOGIC ---

            

            valid_rows_to_insert.append({

                'date': appt_date,

                'patient': patient,

                'service_obj': service_obj, # Store obj

                'doctor_obj': doc_obj,      # Store obj

                'clinic_obj': clinic_obj,   # Store obj

                'manager_id': manager_id,

                'contract': str(get_val('contract') or '').strip(),

                'cost': cost, # Catalog price (Main + Additional)

                'discount_raw': get_val('discount'), # Raw discount

                'payment_raw': str(get_val('payment') or '').strip(),

                'is_child_raw': get_val('is_child'),

                'quantity_raw': get_val('quantity'),

                'quantity_raw': get_val('quantity'),

                'add_service_objs': add_service_objs, # List of objects

                'add_services_raw': add_services_str, # For legacy comment if needed?

                'comment_raw': str(get_val('comment') or '').strip()

            })

            unique_dates_to_clean.add(appt_date)

            

        if not valid_rows_to_insert:

            msg = 'Не найдено валидных строк.'

            if warnings: msg += f" Ошибки ({len(warnings)}): " + "; ".join(warnings[:5]) + "..."

            flash(msg, 'warning')

            return redirect(url_for('admin.additional'))



        if delete_old and unique_dates_to_clean:

             # Fetch IDs to delete dependencies first (query.delete bypasses cascade)

             appts_to_delete = db.session.query(Appointment.id).filter(

                 Appointment.center_id == int(center_id),

                 Appointment.date.in_(unique_dates_to_clean)

             ).all()

             

             appt_ids = [a[0] for a in appts_to_delete]

             

             if appt_ids:

                 # Delete associations manually (using models)

                 # AppointmentAdditionalService (Additional Services)

                 AppointmentAdditionalService.query.filter(

                     AppointmentAdditionalService.appointment_id.in_(appt_ids)

                 ).delete(synchronize_session=False)



                 # AppointmentService (Main Services)

                 AppointmentService.query.filter(

                     AppointmentService.appointment_id.in_(appt_ids)

                 ).delete(synchronize_session=False)



                 # History

                 AppointmentHistory.query.filter(

                     AppointmentHistory.appointment_id.in_(appt_ids)

                 ).delete(synchronize_session=False)



                 # Finally delete Appointments

                 Appointment.query.filter(

                     Appointment.id.in_(appt_ids)

                 ).delete(synchronize_session=False)



        created_count = 0

        current_center_id = int(center_id)

        current_user_id = current_user.id

        

        for data in valid_rows_to_insert:

            # Boolean

            is_child = False

            raw_child = data['is_child_raw']

            if raw_child:

                if isinstance(raw_child, bool): is_child = raw_child

                else:

                    s = str(raw_child).lower()

                    if s in ['true', 'yes', 'да', '1', '+']: is_child = True



            # Payment Method

            pm_name = data['payment_raw']

            pm_id = None

            if pm_name:

                pm = PaymentMethod.query.filter(PaymentMethod.name.ilike(pm_name)).first()

                if pm: pm_id = pm.id

                else:

                     pm = PaymentMethod(name=pm_name)

                     db.session.add(pm)

                     db.session.flush()

                     pm_id = pm.id

            

            # Discount

            discount = 0.0

            if data['discount_raw'] is not None:

                if isinstance(data['discount_raw'], (int, float)):

                    discount = float(data['discount_raw'])

                else:

                    d_str = str(data['discount_raw']).replace(u'\xa0', '').replace(' ', '').replace(',', '.')

                    try: discount = float(d_str)

                    except: discount = 0.0

                

            # Quantity parsed above (excel_qty)

            

            # Legacy Quantity (Main Service) set to 1

            main_qty = 1



            appt = Appointment(

                center_id=current_center_id,

                date=data['date'],

                time="09:00",

                patient_name=data['patient'],

                service=data['service_obj'].name if data['service_obj'] else "Доп. услуги", # Fallback name

                # Link M2M service

                doctor=data['doctor_obj'].name,

                doctor_id=data['doctor_obj'].id,

                manager_id=data['manager_id'], # Auto-assigned

                clinic_id=data['clinic_obj'].id if data['clinic_obj'] else None,

                cost=float(data['cost']),

                discount=discount,

                payment_method_id=pm_id,

                is_child=is_child,

                contract_number=data['contract'],

                quantity=main_qty, # Always 1 for main service

                author_id=current_user_id,

                comment=data['comment_raw']

            )

            

            # Add service to M2M relationship (Force Quantity 1 for Main)

            if data['service_obj']:

                appt.service_associations.append(AppointmentService(service=data['service_obj'], quantity=1))

            

            # Add Additional Services to M2M (Use Excel Quantity)

            if data['add_service_objs']:

                for ads in data['add_service_objs']:

                    appt.additional_service_associations.append(AppointmentAdditionalService(additional_service=ads, quantity=excel_qty))

            

            # Legacy comment? Maybe keep strict? 

            # User said "map column... cost substituted". 

            # If mapped, no need for raw text in comment unless unmapped content exists (which we skip now).

            # So we can remove the comment assignment or keep it if strictly necessary. 

            # Let's keep a cleaner comment:

            if data['add_services_raw']:

                 # If we mapped them, maybe we don't need "Доп: ..." in comment?

                 # But sticking to previous behavior is safer, let's just log it if needed.

                 # Actually, better to NOT duplicate logic in comment if it's real data now.

                 pass 

            

            db.session.add(appt)

            created_count += 1

            

        db.session.commit()

        

        success_msg = f'Успешно импортировано: {created_count}.'

        if warnings:

            # flash works with categories. detailed warning list might be too long.

            # let's show top 5 warnings.

            warn_msg = f" Пропущено строк: {len(warnings)}. Примеры: " + "; ".join(warnings[:5])

            flash(success_msg + warn_msg, 'warning') # Use warning color to draw attention

        else:

            flash(success_msg, 'success')



    except Exception as e:

        db.session.rollback()

        flash(f'Ошибка импорта: {str(e)}', 'error')



    return redirect(url_for('admin.additional'))



@admin.route('/managers/add', methods=['POST'])

def add_manager():

    name = request.form.get('name')

    if not name:

        flash('Имя менеджера обязательно', 'error')

        return redirect(url_for('admin.additional'))

    

    manager = Manager(name=name)

    db.session.add(manager)

    db.session.commit()

    flash(f'Менеджер {name} добавлен', 'success')

    return redirect(url_for('admin.additional'))



@admin.route('/managers/<int:id>/update', methods=['POST'])

def update_manager(id):

    manager = Manager.query.get_or_404(id)

    manager.name = request.form.get('name')

    db.session.commit()

    flash(f'Менеджер {manager.name} обновлен', 'success')

    return redirect(url_for('admin.additional'))



@admin.route('/managers/<int:id>/delete', methods=['POST'])

def delete_manager(id):

    manager = Manager.query.get_or_404(id)

    db.session.delete(manager)

    db.session.commit()

    flash('Менеджер удален', 'success')

    return redirect(url_for('admin.additional'))



# --- Payment Methods Management ---



@admin.route('/payment_methods/add', methods=['POST'])

def add_payment_method():

    name = request.form.get('name')

    if not name:

        flash('Название способа оплаты обязательно', 'error')

        return redirect(url_for('admin.additional'))

    

    method = PaymentMethod(name=name)

    db.session.add(method)

    db.session.commit()

    flash(f'Способ оплаты {name} добавлен', 'success')

    return redirect(url_for('admin.additional'))



@admin.route('/payment_methods/<int:id>/update', methods=['POST'])

def update_payment_method(id):

    method = PaymentMethod.query.get_or_404(id)

    method.name = request.form.get('name')

    db.session.commit()

    flash(f'Способ оплаты {method.name} обновлен', 'success')

    return redirect(url_for('admin.additional'))



@admin.route('/payment_methods/<int:id>/delete', methods=['POST'])

def delete_payment_method(id):

    method = PaymentMethod.query.get_or_404(id)

    db.session.delete(method)

    db.session.commit()

    flash('Способ оплаты удален', 'success')

    return redirect(url_for('admin.additional'))







@admin.route('/')

def index():

    return redirect(url_for('admin.users'))



@admin.route('/chat/settings', methods=['POST'])

@login_required

def update_chat_settings():

    if 'chat_image' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.additional'))

    

    file = request.files['chat_image']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.additional'))

        

    if file:

        filename = secure_filename(file.filename)

        # Save to static/uploads/chat_icons or similar

        upload_dir = os.path.join(current_app.static_folder, 'uploads', 'chat')

        os.makedirs(upload_dir, exist_ok=True)

        

        file_path = os.path.join(upload_dir, filename)

        file.save(file_path)

        

        # Save relative path to DB

        relative_path = f"uploads/chat/{filename}"

        

        setting = GlobalSetting.query.get('chat_image')

        if not setting:

            setting = GlobalSetting(key='chat_image')

            db.session.add(setting)

        

        setting.value = relative_path

        db.session.commit()

        

        flash('Иконка чата обновлена', 'success')

        

    return redirect(url_for('admin.additional'))


@admin.route('/viewer/settings', methods=['POST'])
@login_required
def update_viewer_settings():
    guac_url = request.form.get('guacamole_base_url')
    if guac_url:
        setting = GlobalSetting.query.get('guacamole_base_url')
        if not setting:
            setting = GlobalSetting(key='guacamole_base_url')
            db.session.add(setting)
        setting.value = guac_url
        db.session.commit()
        flash('Настройки просмотрщика обновлены', 'success')
    return redirect(url_for('admin.additional'))



@admin.route('/stamp/upload', methods=['POST'])

@login_required

def upload_stamp_image():

    if 'stamp_image' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.additional'))

    

    file = request.files['stamp_image']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.additional'))

        

    if file:

        filename = secure_filename(file.filename)

        # Save to static/uploads/stamps

        upload_dir = os.path.join(current_app.static_folder, 'uploads', 'stamps')

        os.makedirs(upload_dir, exist_ok=True)

        

        file_path = os.path.join(upload_dir, filename)

        file.save(file_path)

        

        # Save relative path to DB

        relative_path = f"uploads/stamps/{filename}"

        

        setting = GlobalSetting.query.get('stamp_image')

        if not setting:

            setting = GlobalSetting(key='stamp_image')

            db.session.add(setting)

        

        setting.value = relative_path

        db.session.commit()

        

        flash('Изображение печати обновлено', 'success')

        

    return redirect(url_for('admin.additional'))


# ========== Medical Certificate Generator ==========

@admin.route('/stamp-tool/patients', methods=['GET'])
@login_required
def get_patients_for_certificate():
    """Get list of recent patients with appointment data"""
    try:
        # Get recent appointments (last 3 months) with unique patients
        from datetime import date
        three_months_ago = date.today() - timedelta(days=90)
        
        appointments = Appointment.query.filter(
            Appointment.date >= three_months_ago
        ).order_by(Appointment.date.desc()).limit(500).all()
        
        # Create unique patient list with their data
        patients_data = []
        seen_names = set()
        
        for apt in appointments:
            if apt.patient_name not in seen_names:
                patients_data.append({
                    'id': apt.id,
                    'patient_name': apt.patient_name,
                    'cost': apt.cost,
                    'date': apt.date.isoformat()
                })
                seen_names.add(apt.patient_name)
        
        return jsonify({'success': True, 'patients': patients_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin.route('/stamp-tool/certificate/generate', methods=['POST'])
@login_required
@csrf.exempt
def generate_certificate():
    """Generate medical certificate JPG from template"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        from datetime import date
        
        data = request.get_json()
        
        # Extract data
        appointment_id = data.get('appointment_id')
        patient_name = data.get('patient_name')
        inn = data.get('inn', '')
        birth_date_str = data.get('birth_date')
        doc_series = data.get('doc_series', '')
        doc_number = data.get('doc_number', '')
        doc_issue_date_str = data.get('doc_issue_date')
        amount = float(data.get('amount', 0))
        
        # Parse dates
        birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date() if birth_date_str else None
        doc_issue_date = datetime.strptime(doc_issue_date_str, '%Y-%m-%d').date() if doc_issue_date_str else None
        
        # Split patient name into parts (Фамилия Имя Отчество)
        name_parts = patient_name.split()
        surname = name_parts[0] if len(name_parts) > 0 else ''
        name = name_parts[1] if len(name_parts) > 1 else ''
        patronymic = name_parts[2] if len(name_parts) > 2 else ''
        
        # Load template
        template_path = os.path.join(current_app.root_path, '..', 'orbital logo files', 'Без заказчика.png')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Template not found'}), 404
        
        img = Image.open(template_path).convert('RGB')
        draw = ImageDraw.Draw(img)
        
        # Font setup (try different fonts, fallback to default)
        try:
            font = ImageFont.truetype('C:\\Windows\\Fonts\\arial.ttf', 28)
            font_small = ImageFont.truetype('C:\\Windows\\Fonts\\arial.ttf', 24)
        except:
            font = ImageFont.load_default()
            font_small = font
        
        # Draw fields on template (coordinates to be adjusted based on actual template)
        # These are placeholder coordinates - need to be calibrated
        draw.text((200, 250), surname, fill='black', font=font)  # Фамилия
        draw.text((200, 300), name, fill='black', font=font)  # Имя
        draw.text((200, 350), patronymic, fill='black', font=font)  # Отчество
        
        if inn:
            draw.text((600, 200), inn, fill='black', font=font_small)  # ИНН
        
        if birth_date:
            draw.text((400, 400), birth_date.strftime('%d.%m.%Y'), fill='black', font=font_small)  # Дата рождения
        
        draw.text((200, 450), doc_series, fill='black', font=font_small)  # Серия
        draw.text((350, 450), doc_number, fill='black', font=font_small)  # Номер
        
        if doc_issue_date:
            draw.text((550, 450), doc_issue_date.strftime('%d.%m.%Y'), fill='black', font=font_small)  # Дата выдачи
        
        # Auto-filled fields
        draw.text((700, 500), '21', fill='black', font=font)  # Код документа
        draw.text((850, 550), '1', fill='black', font=font)  # Налогоплательщик = пациент
        draw.text((400, 600), f'{amount:.2f}', fill='black', font=font)  # Сумма
        draw.text((700, 650), date.today().strftime('%d.%m.%Y'), fill='black', font=font_small)  # Сегодняшняя дата
        
        # Overlay stamp at bottom
        stamp_setting = GlobalSetting.query.get('stamp_image')
        if stamp_setting and stamp_setting.value:
            stamp_path = os.path.join(current_app.static_folder, stamp_setting.value)
            if os.path.exists(stamp_path):
                stamp = Image.open(stamp_path).convert('RGBA')
                # Resize stamp if needed
                stamp = stamp.resize((200, 200), Image.Resampling.LANCZOS)
                # Paste at bottom (adjust coordinates)
                img.paste(stamp, (300, img.height - 250), stamp)
        
        # Save certificate
        cert_dir = os.path.join(current_app.static_folder, 'uploads', 'certificates')
        os.makedirs(cert_dir, exist_ok=True)
        
        filename = f'cert_{date.today().strftime("%Y_%m_%d")}_{int(datetime.now().timestamp())}.jpg'
        filepath = os.path.join(cert_dir, filename)
        
        img.save(filepath, 'JPEG', quality=95)
        
        # Save to database
        cert = MedicalCertificate(
            appointment_id=appointment_id,
            patient_name=patient_name,
            inn=inn,
            birth_date=birth_date,
            doc_series=doc_series,
            doc_number=doc_number,
            doc_issue_date=doc_issue_date,
            amount=amount,
            filename=filename,
            created_by_id=current_user.id
        )
        
        db.session.add(cert)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'certificate_id': cert.id,
            'filename': filename,
            'download_url': url_for('admin.download_certificate', cert_id=cert.id)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Generation failed: {str(e)}'}), 500


@admin.route('/stamp-tool/certificate/<int:cert_id>/download')
@login_required
def download_certificate(cert_id):
    """Download generated certificate"""
    cert = MedicalCertificate.query.get_or_404(cert_id)
    filepath = os.path.join(current_app.static_folder, 'uploads', 'certificates', cert.filename)
    
    if not os.path.exists(filepath):
        abort(404)
    
    from flask import send_file
    return send_file(filepath, as_attachment=True, download_name=f'certificate_{cert.patient_name}_{cert.id}.jpg')


@admin.route('/stamp-tool/certificates', methods=['GET'])
@login_required
def list_certificates():
    """List generated certificates"""
    certificates = MedicalCertificate.query.order_by(
        MedicalCertificate.generated_at.desc()
    ).limit(50).all()
    
    return jsonify({
        'success': True,
        'certificates': [cert.to_dict() for cert in certificates]
    })


def cleanup_old_certificates():
    """Cleanup certificates older than 30 days"""
    try:
        cutoff = datetime.utcnow() - timedelta(days=30)
        old_certs = MedicalCertificate.query.filter(
            MedicalCertificate.generated_at < cutoff
        ).all()
        
        for cert in old_certs:
            # Delete file
            filepath = os.path.join(current_app.static_folder, 'uploads', 'certificates', cert.filename)
            if os.path.exists(filepath):
                os.remove(filepath)
            # Delete DB record
            db.session.delete(cert)
        
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        print(f"Certificate cleanup error: {e}")
        return False


# ========== Monitoring Metrics Collection ==========

# Simple in-memory cache for statistics (5-minute expiration)
_stats_cache = {'data': None, 'timestamp': None}

def get_cached_statistics():
    """Get system statistics with 5-minute cache"""
    now = datetime.utcnow()
    
    # Check if cache is valid
    if (_stats_cache['timestamp'] and 
        now - _stats_cache['timestamp'] < timedelta(minutes=5)):
        return _stats_cache['data']
    
    # Recalculate statistics
    stats = {
        'users_count': User.query.count(),
        'appointments_count': Appointment.query.count(),
        'journal_entries_count': AppointmentHistory.query.count(),
        'doctors_count': Doctor.query.count(),
        'clinics_count': Location.query.filter_by(type='center').count() + 
                        Location.query.filter(Location.type != 'center', Location.type != 'city').count(),
        'organizations_count': Organization.query.count(),
        'services_count': Service.query.count(),
        'locations_count': Location.query.count()
    }
    
    # Update cache
    _stats_cache['data'] = stats
    _stats_cache['timestamp'] = now
    
    return stats

def collect_system_metrics():
    """Collect all system metrics and save to database"""
    try:
        # Get current statistics
        stats = get_cached_statistics()
        
        # Get system resource usage
        cpu_percent = psutil.cpu_percent(interval=1)
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        
        # Create new metrics record
        metrics = SystemMetrics(
            timestamp=datetime.utcnow(),
            disk_total_gb=round(disk.total / (1024**3), 2),
            disk_used_gb=round(disk.used / (1024**3), 2),
            disk_percent=disk.percent,
            users_count=stats['users_count'],
            appointments_count=stats['appointments_count'],
            journal_entries_count=stats['journal_entries_count'],
            doctors_count=stats['doctors_count'],
            clinics_count=stats['clinics_count'],
            organizations_count=stats['organizations_count'],
            services_count=stats['services_count'],
            cpu_percent=cpu_percent,
            ram_percent=mem.percent
        )
        
        db.session.add(metrics)
        db.session.commit()
        
        # Cleanup old metrics (keep 30 days)
        cutoff_date = datetime.utcnow() - timedelta(days=30)
        SystemMetrics.query.filter(SystemMetrics.timestamp < cutoff_date).delete()
        db.session.commit()
        
        return True
    except Exception as e:
        db.session.rollback()
        print(f"Error collecting metrics: {e}")
        return False


@admin.route('/monitoring')

@login_required

def monitoring():

    # Database Status

    try:

        db.session.execute(db.text('SELECT 1'))

        db_status = True

    except Exception:

        db_status = False

    

    # Bot Status

    bot_configured = bool(telegram_bot.token and telegram_bot.chat_id)

    

    # Current System Stats

    cpu_percent = psutil.cpu_percent(interval=None) # Non-blocking

    

    mem = psutil.virtual_memory()

    ram_total_gb = round(mem.total / (1024**3), 2)

    ram_used_gb = round(mem.used / (1024**3), 2)

    ram_percent = mem.percent

    

    disk = psutil.disk_usage('/')

    disk_total_gb = round(disk.total / (1024**3), 2)

    disk_free_gb = round(disk.free / (1024**3), 2)

    disk_percent = disk.percent
    
    # Get cached statistics
    stats = get_cached_statistics()
    
    # Get historical metrics for graph (last 7 days)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    historical_metrics = SystemMetrics.query.filter(
        SystemMetrics.timestamp >= seven_days_ago
    ).order_by(SystemMetrics.timestamp.asc()).all()
    
    # Format graph data
    graph_data = {
        'labels': [m.timestamp.strftime('%d.%m') for m in historical_metrics],
        'disk_used': [m.disk_used_gb for m in historical_metrics],
        'disk_percent': [m.disk_percent for m in historical_metrics]
    }

    

    return render_template('admin_monitoring.html', 

                           db_status=db_status,

                           bot_configured=bot_configured,

                           cpu_percent=cpu_percent,

                           ram_total_gb=ram_total_gb,

                           ram_used_gb=ram_used_gb,

                           ram_percent=ram_percent,

                           disk_total_gb=disk_total_gb,

                           disk_free_gb=disk_free_gb,

                           disk_percent=disk_percent,
                           
                           # Statistics
                           stats=stats,
                           
                           # Graph data
                           graph_data=graph_data

                           )


@admin.route('/monitoring/refresh', methods=['POST'])
@login_required
def monitoring_refresh():
    """Manually refresh statistics cache and collect metrics"""
    # Clear cache
    global _stats_cache
    _stats_cache = {'data': None, 'timestamp': None}
    
    # Collect new metrics snapshot
    success = collect_system_metrics()
    
    return jsonify({'success': success})


@admin.route('/monitoring/test-message', methods=['POST'])

@login_required

def test_telegram_message():

    success = telegram_bot.send_message("✅ <b>ТЕСТОВОЕ СООБЩЕНИЕ</b>\n\nБот работает корректно!")

    return jsonify({'success': success})



@admin.route('/users/add', methods=['POST'])

def add_user():

    username = request.form.get('username')

    email = request.form.get('email')

    password = request.form.get('password')

    role = request.form.get('role')

    organization_id = request.form.get('organization_id')

    city_id = request.form.get('city_id')

    center_id = request.form.get('center_id')
    clinic_id = request.form.get('clinic_id')
    doctor_id = request.form.get('doctor_id')


    if User.query.filter_by(username=username).first():

        flash('Пользователь с таким именем уже существует', 'error')

        return redirect(url_for('admin.users'))

    if User.query.filter_by(email=email).first():

        flash('Пользователь с таким email уже существует', 'error')

        return redirect(url_for('admin.users'))



    new_user = User(
        username=username,
        email=email,
        role=role,
        is_confirmed=False  # Require email confirmation
    )

    new_user.password_hash = generate_password_hash(password)
    
    if organization_id:
        new_user.organization_id = int(organization_id)
    if city_id:
        new_user.city_id = int(city_id)
    if center_id:
        new_user.center_id = int(center_id)
    if clinic_id:
        new_user.clinic_id = int(clinic_id)
    if doctor_id:
        new_user.doctor_id = int(doctor_id)

    db.session.add(new_user)
    db.session.commit()
    
    # Send Confirmation Email
    try:
        from app.extensions import mail
        from flask_mail import Message
        from itsdangerous import URLSafeTimedSerializer


        ts = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
        token = ts.dumps(email, salt='email-confirm-key')
        confirm_url = url_for('auth.confirm_email', token=token, _external=True)

        msg = Message('Подтвердите ваш email', recipients=[email])
        msg.body = f'Здравствуйте! Пожалуйста, перейдите по следующей ссылке для подтверждения вашего аккаунта: {confirm_url}'
        
        mail.send(msg)
        flash(f'Пользователь {username} создан. Письмо с подтверждением отправлено на {email}', 'success')
    except Exception as e:
        flash(f'Пользователь создан, но не удалось отправить письмо подтверждения: {str(e)}', 'warning')

    return redirect(url_for('admin.users'))

@admin.route('/users/edit/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if current_user.role not in ['superadmin', 'admin']:
        flash('У вас нет прав для выполнения этого действия', 'error')
        return redirect(url_for('admin.users'))
        
    user = User.query.get_or_404(user_id)
    
    # Optional: Prevent editing superadmins by non-superadmins, etc. 
    # For now, simplistic check:
    if user.role == 'superadmin' and current_user.role != 'superadmin':
         flash('Вы не можете редактировать суперадминистратора', 'error')
         return redirect(url_for('admin.users'))

    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    organization_id = request.form.get('organization_id')
    city_id = request.form.get('city_id')
    center_id = request.form.get('center_id')
    clinic_id = request.form.get('clinic_id')
    doctor_id = request.form.get('doctor_id')

    # Update fields
    if username: user.username = username
    if email: user.email = email
    if role: user.role = role
    
    # Password update only if provided
    if password and password.strip():
        user.password_hash = generate_password_hash(password)

    # Relations
    # Handle "None" or empty strings
    if organization_id: user.organization_id = int(organization_id)
    else: user.organization_id = None
        
    if city_id: user.city_id = int(city_id)
    else: user.city_id = None
        
    if center_id: user.center_id = int(center_id)
    else: user.center_id = None

    if clinic_id: user.clinic_id = int(clinic_id)
    else: user.clinic_id = None

    if doctor_id: user.doctor_id = int(doctor_id)
    else: user.doctor_id = None

    try:
        db.session.commit()
        flash(f'Пользователь {user.username} успешно обновлен', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка обновления: {str(e)}', 'error')

    return redirect(url_for('admin.users'))



# --- User Management ---



@admin.route('/users')

def users():

    role_filter = request.args.get('role')

    query = User.query

    

    if role_filter:

        query = query.filter_by(role=role_filter)

        

    users = query.order_by(User.username.asc()).all()

    centers = Location.query.filter_by(type='center').all()

    cities = Location.query.filter_by(type='city').all()

    organizations = Organization.query.all()
    
    clinics = Clinic.query.order_by(Clinic.name).all()

    doctors = Doctor.query.order_by(Doctor.name).all()

    return render_template('admin_users.html', users=users, centers=centers, cities=cities, organizations=organizations, clinics=clinics, doctors=doctors, current_role=role_filter)



@admin.route('/users/<int:user_id>/role', methods=['POST'])

def update_user_role(user_id):

    user = User.query.get_or_404(user_id)

    new_role = request.form.get('role')

    

    role_map = {
        'superadmin': 'Суперадмин',
        'admin': 'Админ',
        'org': 'Организация',
        'lab_tech': 'Лаборант',
        'doctor': 'Врач'
    }

    
    if new_role not in ['superadmin', 'admin', 'org', 'lab_tech', 'doctor']:
        flash('Недопустимая роль', 'error')

    else:

        user.role = new_role

        db.session.commit()

        role_name = role_map.get(new_role, new_role)

        flash(f'Роль пользователя {user.username} обновлена на "{role_name}"', 'success')

    

    return redirect(url_for('admin.users'))



@admin.route('/users/<int:user_id>/center', methods=['POST'])

def update_user_center(user_id):

    user = User.query.get_or_404(user_id)

    center_id = request.form.get('center_id')

    

    if center_id:

        user.center_id = int(center_id)

    else:

        user.center_id = None

        

    db.session.commit()

    flash(f'Центр пользователя {user.username} обновлен', 'success')

    return redirect(url_for('admin.users'))



@admin.route('/users/<int:user_id>/block', methods=['POST'])

def block_user(user_id):

    user = User.query.get_or_404(user_id)

    if user.role == 'admin':

        flash('Нельзя заблокировать администратора', 'error')

    else:

        user.is_blocked = True

        db.session.commit()

        flash(f'Пользователь {user.username} заблокирован', 'success')

    return redirect(url_for('admin.users'))



@admin.route('/users/<int:user_id>/unblock', methods=['POST'])

def unblock_user(user_id):

    user = User.query.get_or_404(user_id)

    user.is_blocked = False

    db.session.commit()

    flash(f'Пользователь {user.username} разблокирован', 'success')

    return redirect(url_for('admin.users'))



@admin.route('/users/<int:user_id>/confirm', methods=['POST'])

def confirm_user(user_id):

    user = User.query.get_or_404(user_id)

    user.is_confirmed = True

    db.session.commit()

    flash(f'Пользователь {user.username} подтвержден', 'success')

    return redirect(url_for('admin.users'))



# --- Location Management ---



@admin.route('/locations')

def locations():

    cities = Location.query.filter_by(type='city').all()

    return render_template('admin_locations.html', cities=cities)



@admin.route('/locations/add', methods=['POST'])

def add_location():

    name = request.form.get('name')

    type = request.form.get('type')

    parent_id = request.form.get('parent_id')

    color = request.form.get('color')



    if not name or type not in ['city', 'center']:

        return jsonify({'error': 'Неверные данные'}), 400



    loc = Location(name=name, type=type)

    if parent_id:

        loc.parent_id = int(parent_id)

    if color:

        loc.color = color

    

    db.session.add(loc)

    db.session.commit()

    return jsonify(loc.to_dict()), 201



@admin.route('/locations/<int:id>/edit', methods=['PUT'])

def edit_location(id):

    loc = Location.query.get_or_404(id)

    data = request.get_json()

    

    if 'name' in data:

        loc.name = data['name']

    if 'color' in data:

        loc.color = data['color']

    

    db.session.commit()

    return jsonify(loc.to_dict())



@admin.route('/locations/<int:id>/delete', methods=['DELETE'])

def delete_location(id):

    loc = Location.query.get_or_404(id)

    # Optional: Check for children or related data before deleting

    db.session.delete(loc)

    db.session.commit()

    return '', 204



# --- Doctor Management ---



@admin.route('/doctors')
def doctors():
    doctors = Doctor.query.order_by(Doctor.name).all()
    managers = Manager.query.all()
    
    # Get active bonus period for columns count
    today = date.today()
    period = BonusPeriod.query.filter(
        BonusPeriod.start_date <= today
    ).filter(
        (BonusPeriod.end_date >= today) | (BonusPeriod.end_date == None)
    ).order_by(BonusPeriod.start_date.desc()).first()
    
    bonus_cols = period.columns if period else 0
    
    return render_template('admin_doctors.html', doctors=doctors, managers=managers, bonus_cols=bonus_cols)



@admin.route('/doctors/add', methods=['POST'])
def add_doctor():
    name = request.form.get('name')
    specialization = request.form.get('specialization')
    manager = request.form.get('manager')
    bonus_type = request.form.get('bonus_type')
    
    if not name:
        flash('Имя врача обязательно', 'error')
        return redirect(url_for('admin.doctors'))
        
    doctor = Doctor(name=name, specialization=specialization, manager=manager)
    if bonus_type:
        doctor.bonus_type = int(bonus_type)
        
    db.session.add(doctor)
    db.session.commit()
    flash(f'Врач {name} добавлен', 'success')
    return redirect(url_for('admin.doctors'))

@admin.route('/doctors/<int:id>/update', methods=['POST'])
def update_doctor(id):
    doctor = Doctor.query.get_or_404(id)
    doctor.name = request.form.get('name')
    doctor.specialization = request.form.get('specialization')
    doctor.manager = request.form.get('manager')
    
    bonus_type = request.form.get('bonus_type')
    if bonus_type:
        doctor.bonus_type = int(bonus_type)
    else:
        doctor.bonus_type = None

    db.session.commit()
    flash(f'Данные врача {doctor.name} обновлены', 'success')

    return redirect(url_for('admin.doctors'))



@admin.route('/doctors/<int:id>/delete', methods=['POST'])

def delete_doctor(id):

    doctor = Doctor.query.get_or_404(id)

    db.session.delete(doctor)

    db.session.commit()

    flash('Врач удален', 'success')

    return redirect(url_for('admin.doctors'))



@admin.route('/doctors/import', methods=['POST'])

def import_doctors():

    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.doctors'))

    

    file = request.files['file']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.doctors'))



    if file:

        filename = secure_filename(file.filename)

        doctors_created = 0

        errors = 0

        

        try:

            # Determine file type

            if filename.endswith('.csv'):

                # Handle CSV

                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)

                csv_input = csv.reader(stream)

                # Assume header: Name, Specialization, Manager

                # Or just skip first row if it looks like header?

                # Let's assume with header if "Name" is in first row, or just simple loop.

                # Simplest: Assume Header exists.

                header = next(csv_input, None)

                

                for row in csv_input:

                    if len(row) >= 1: # At least name

                        name = row[0].strip()

                        if not name: continue

                        specialization = row[1].strip() if len(row) > 1 else None

                        manager = row[2].strip() if len(row) > 2 else None

                        

                        # Check exist

                        if not Doctor.query.filter_by(name=name).first():

                            doc = Doctor(name=name, specialization=specialization, manager=manager)

                            db.session.add(doc)

                            doctors_created += 1

                        

            elif filename.endswith('.xlsx'):

                # Handle Excel

                wb = openpyxl.load_workbook(file)

                sheet = wb.active

                

                # Iterate rows, skipping header (min_row=2)

                for row in sheet.iter_rows(min_row=2, values_only=True):

                    if not row or not row[0]: continue

                    name = str(row[0]).strip()

                    specialization = str(row[1]).strip() if len(row) > 1 and row[1] else None

                    manager = str(row[2]).strip() if len(row) > 2 and row[2] else None

                    

                    if not Doctor.query.filter_by(name=name).first():

                        doc = Doctor(name=name, specialization=specialization, manager=manager)

                        db.session.add(doc)

                        doctors_created += 1



            else:

                 flash('Неподдерживаемый формат файла. Используйте CSV или XLSX.', 'error')

                 return redirect(url_for('admin.doctors'))



            db.session.commit()

            flash(f'Импортировано врачей: {doctors_created}', 'success')



        except Exception as e:

            db.session.rollback()

            flash(f'Ошибка импорта: {str(e)}', 'error')

            

    return redirect(url_for('admin.doctors'))



# --- Service Management ---



@admin.route('/services')
def services():
    services = Service.query.order_by(Service.name.asc()).all()
    return render_template('admin_services.html', services=services)



@admin.route('/services/add', methods=['POST'])

def add_service():

    name = request.form.get('name')

    price = request.form.get('price')

    

    if not name or not price:

        flash('Все поля обязательны', 'error')

        return redirect(url_for('admin.services'))

        

    service = Service(name=name, price=float(price))

    db.session.add(service)

    db.session.commit()

    flash(f'Услуга {name} добавлена', 'success')

    return redirect(url_for('admin.services'))



@admin.route('/services/<int:id>/update', methods=['POST'])

def update_service(id):

    service = Service.query.get_or_404(id)

    service.name = request.form.get('name')

    service.price = float(request.form.get('price'))

    db.session.commit()

    flash(f'Услуга {service.name} обновлена', 'success')

    return redirect(url_for('admin.services'))



@admin.route('/services/<int:id>/delete', methods=['POST'])

def delete_service(id):

    service = Service.query.get_or_404(id)
    db.session.delete(service)
    db.session.commit()
    flash('Услуга удалена', 'success')
    return redirect(url_for('admin.services'))

@admin.route('/services/<int:id>/toggle_visibility', methods=['POST'])
@login_required
def toggle_service_visibility(id):
    if current_user.role != 'superadmin':
        abort(403)
        
    service = Service.query.get_or_404(id)
    service.is_hidden = not service.is_hidden
    db.session.commit()
    
    status = "скрыта" if service.is_hidden else "видима"
    flash(f'Услуга "{service.name}" теперь {status}', 'success')
    return redirect(url_for('admin.services'))


@admin.route('/services/import', methods=['POST'])

def import_services():

    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.services'))

    

    file = request.files['file']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.services'))



    if file:

        filename = secure_filename(file.filename)

        services_created = 0

        

        try:

            # Determine file type

            if filename.endswith('.csv'):

                # Handle CSV

                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)

                csv_input = csv.reader(stream)

                header = next(csv_input, None)

                

                for row in csv_input:

                    if len(row) >= 2: # Name, Price required

                        name = row[0].strip()

                        if not name: continue

                        try:

                            price = float(row[1].strip())

                        except ValueError:

                            continue

                            

                        # Check exist

                        if not Service.query.filter_by(name=name).first():

                            service = Service(name=name, price=price)

                            db.session.add(service)

                            services_created += 1

                        

            elif filename.endswith('.xlsx'):

                # Handle Excel

                wb = openpyxl.load_workbook(file)

                sheet = wb.active

                

                # Iterate rows, skipping header (min_row=2)

                for row in sheet.iter_rows(min_row=2, values_only=True):

                    if not row or not row[0]: continue

                    name = str(row[0]).strip()

                    try:

                        price = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0

                    except (ValueError, TypeError):

                        continue

                    

                    if not Service.query.filter_by(name=name).first():

                        service = Service(name=name, price=price)

                        db.session.add(service)

                        services_created += 1



            else:

                 flash('Неподдерживаемый формат файла. Используйте CSV или XLSX.', 'error')

                 return redirect(url_for('admin.services'))



            db.session.commit()

            flash(f'Импортировано услуг: {services_created}', 'success')



        except Exception as e:

            db.session.rollback()

            flash(f'Ошибка импорта: {str(e)}', 'error')

            

    return redirect(url_for('admin.services'))



@admin.route('/services/<int:id>/prices')

def service_prices(id):

    service = Service.query.get_or_404(id)

    # Sort prices by start_date desc

    prices = ServicePrice.query.filter_by(service_id=id).order_by(ServicePrice.start_date.desc()).all()

    return render_template('admin_service_prices.html', service=service, prices=prices)



@admin.route('/services/<int:id>/prices/add', methods=['POST'])

def add_service_price(id):

    service = Service.query.get_or_404(id)

    price_val = request.form.get('price')

    start_date_str = request.form.get('start_date')

    end_date_str = request.form.get('end_date')

    

    if not price_val or not start_date_str:

        flash('Цена и Дата начала обязательны', 'error')

        return redirect(url_for('admin.service_prices', id=id))

        

    try:

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

        end_date = None

        if end_date_str:

             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

             

        # Basic Validation: Start date must be before end date

        if end_date and start_date > end_date:

            flash('Дата начала не может быть позже даты окончания', 'error')

            return redirect(url_for('admin.service_prices', id=id))



        new_price = ServicePrice(

            service_id=service.id,

            price=float(price_val),

            start_date=start_date,

            end_date=end_date

        )

        db.session.add(new_price)

        db.session.commit()

        flash('Период действия цены добавлен', 'success')

        

    except ValueError:

        flash('Ошибка формата данных', 'error')

        

    return redirect(url_for('admin.service_prices', id=id))



@admin.route('/services/prices/<int:price_id>/delete', methods=['POST'])

def delete_service_price(price_id):

    price_entry = ServicePrice.query.get_or_404(price_id)

    service_id = price_entry.service_id

    db.session.delete(price_entry)

    db.session.commit()

    flash('Период удален', 'success')

    return redirect(url_for('admin.service_prices', id=service_id))



# --- Additional Service Management ---



@admin.route('/additional_services')

def additional_services():

    services = AdditionalService.query.order_by(AdditionalService.name.asc()).all()

    return render_template('admin_additional_services.html', additional_services=services)



@admin.route('/additional_services/add', methods=['POST'])

def add_additional_service():

    name = request.form.get('name')

    price = request.form.get('price')

    

    if not name or not price:

        flash('Все поля обязательны', 'error')

        return redirect(url_for('admin.additional_services'))

        

    service = AdditionalService(name=name, price=float(price))

    db.session.add(service)

    db.session.commit()

    flash(f'Доп. услуга {name} добавлена', 'success')

    return redirect(url_for('admin.additional_services'))



@admin.route('/additional_services/<int:id>/update', methods=['POST'])

def update_additional_service(id):

    service = AdditionalService.query.get_or_404(id)

    service.name = request.form.get('name')

    # Only update price if provided (for name-only editing)
    price = request.form.get('price')
    if price:
        service.price = float(price)

    db.session.commit()

    flash(f'Доп. услуга {service.name} обновлена', 'success')

    return redirect(url_for('admin.additional_services'))



@admin.route('/additional_services/<int:id>/delete', methods=['POST'])

def delete_additional_service(id):

    service = AdditionalService.query.get_or_404(id)

    db.session.delete(service)

    db.session.commit()

    flash('Доп. услуга удалена', 'success')

    return redirect(url_for('admin.additional_services'))



@admin.route('/additional_services/import', methods=['POST'])

def import_additional_services():

    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.additional_services'))

    

    file = request.files['file']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.additional_services'))



    if file:

        filename = secure_filename(file.filename)

        services_created = 0

        

        try:

            # Determine file type

            if filename.endswith('.csv'):

                # Handle CSV

                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)

                csv_input = csv.reader(stream)

                header = next(csv_input, None)

                

                for row in csv_input:

                    if len(row) >= 2: # Name, Price required

                        name = row[0].strip()

                        if not name: continue

                        try:

                            price = float(row[1].strip())

                        except ValueError:

                            continue

                            

                        # Check exist

                        if not AdditionalService.query.filter_by(name=name).first():

                            service = AdditionalService(name=name, price=price)

                            db.session.add(service)

                            services_created += 1

                        

            elif filename.endswith('.xlsx'):

                # Handle Excel

                wb = openpyxl.load_workbook(file)

                sheet = wb.active

                

                # Iterate rows, skipping header (min_row=2)

                for row in sheet.iter_rows(min_row=2, values_only=True):

                    if not row or not row[0]: continue

                    name = str(row[0]).strip()

                    try:

                        price = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0

                    except (ValueError, TypeError):

                        continue

                    

                    if not AdditionalService.query.filter_by(name=name).first():

                        service = AdditionalService(name=name, price=price)

                        db.session.add(service)

                        services_created += 1



            else:

                 flash('Неподдерживаемый формат файла. Используйте CSV или XLSX.', 'error')

                 return redirect(url_for('admin.additional_services'))



            db.session.commit()

            flash(f'Импортировано доп. услуг: {services_created}', 'success')



        except Exception as e:

            db.session.rollback()

            flash(f'Ошибка импорта: {str(e)}', 'error')

            

    return redirect(url_for('admin.additional_services'))



@admin.route('/additional_services/<int:id>/prices')

def additional_service_prices(id):

    service = AdditionalService.query.get_or_404(id)

    # Sort prices by start_date desc

    prices = AdditionalServicePrice.query.filter_by(additional_service_id=id).order_by(AdditionalServicePrice.start_date.desc()).all()

    return render_template('admin_additional_service_prices.html', service=service, prices=prices)



@admin.route('/additional_services/<int:id>/prices/add', methods=['POST'])

def add_additional_service_price(id):

    service = AdditionalService.query.get_or_404(id)

    price_val = request.form.get('price')

    start_date_str = request.form.get('start_date')

    end_date_str = request.form.get('end_date')

    

    if not price_val or not start_date_str:

        flash('Цена и Дата начала обязательны', 'error')

        return redirect(url_for('admin.additional_service_prices', id=id))

        

    try:

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

        end_date = None

        if end_date_str:

             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

             

        if end_date and start_date > end_date:

            flash('Дата начала не может быть позже даты окончания', 'error')

            return redirect(url_for('admin.additional_service_prices', id=id))



        new_price = AdditionalServicePrice(

            additional_service_id=service.id,

            price=float(price_val),

            start_date=start_date,

            end_date=end_date

        )

        db.session.add(new_price)

        db.session.commit()

        flash('Период действия цены добавлен', 'success')

        

    except ValueError:

        flash('Ошибка формата данных', 'error')

        

    return redirect(url_for('admin.additional_service_prices', id=id))



@admin.route('/additional_services/prices/<int:price_id>/delete', methods=['POST'])

def delete_additional_service_price(price_id):

    price_entry = AdditionalServicePrice.query.get_or_404(price_id)

    service_id = price_entry.additional_service_id

    db.session.delete(price_entry)

    db.session.commit()

    flash('Период удален', 'success')

    return redirect(url_for('admin.additional_service_prices', id=service_id))



    flash('Период удален', 'success')

    return redirect(url_for('admin.service_prices', id=service_id))



# --- Journal Import ---



@admin.route('/journal/import', methods=['POST'])

@login_required

def import_journal():

    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('main.journal'))

    

    file = request.files['file']

    center_id = request.form.get('center_id')

    

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('main.journal'))

        

    if not center_id:

        flash('Не выбран Центр', 'error')

        return redirect(url_for('main.journal'))



    if file:

        filename = secure_filename(file.filename)

        created_count = 0

        try:

             if filename.endswith('.xlsx'):

                wb = openpyxl.load_workbook(file, data_only=True)

                sheet = wb.active

                

                # Iterate rows, skipping header (min_row=2)

                for row in sheet.iter_rows(min_row=2, values_only=True):

                    # Check Row Structure: 

                    # 0: Date, 1: Month, 2: Lab Tech, 3: Contract, 4: Patient, 5: Child (BOOL), 

                    # 6: Doctor, 7: Manager, 8: Clinic, 9: Service, 10: Add Svc, 11: Qty, 

                    # 12: Add Svc Qty, 13: Cost (Text?), 14: Payment, 15: Discount, 16: Sum, 17: Comment

                    

                    if not row or not row[0]: continue

                    

                    # 1. Date

                    date_val = row[0]

                    if isinstance(date_val, datetime):

                        appt_date = date_val.date()

                    else:

                        # Try parse string?

                        try:

                            # Try common formats or just skip

                             appt_date = datetime.strptime(str(date_val), '%Y-%m-%d').date() # fallback

                        except:

                            continue # Skip invalid date

                            

                    # 2. Service

                    service_name = str(row[9]).strip() if len(row) > 9 and row[9] else None

                    if not service_name: continue

                    

                    # Find Service object for price calc

                    service = Service.query.filter(Service.name.ilike(service_name)).first()

                    # If service not found, creating it might be dangerous for price calc. 

                    # But we need it for the record. 

                    # Let's create if missing with 0 price? Or skip?

                    # Plan said: "Search by name. If found, link service_id? Wait, Appointment stores Service Name string primarily in old model, 

                    # but we moved to ID in form. 

                    # Model: service = db.Column(db.String(200)) ... wait, check model.

                    # Model: service = db.Column(db.String(200), nullable=False)

                    # It does NOT have service_id foreign key? 

                    # Checking model... "service = db.Column(db.String(200), nullable=False)"

                    # Ah, so we just store name. BUT we need price.

                    # So we find the Service model to get price.

                    

                    price = 0.0

                    if service:

                        price = service.get_price(appt_date)

                    

                    # 3. Quantity

                    try:

                        qty = int(row[11]) if len(row) > 11 and row[11] else 1

                    except:

                        qty = 1

                        

                    # 4. Discount

                    try:

                        discount = float(row[15]) if len(row) > 15 and row[15] else 0.0

                    except:

                        discount = 0.0

                        

                    # 5. Calculate Cost

                    final_cost = (price * qty) - discount

                    if final_cost < 0: final_cost = 0.0

                    

                    # 6. Patient

                    patient_name = str(row[4]).strip() if len(row) > 4 and row[4] else "Unknown"

                    

                    # 7. Doctor

                    doctor_name = str(row[6]).strip() if len(row) > 6 and row[6] else "Unknown"

                    doctor_obj = Doctor.query.filter(Doctor.name.ilike(doctor_name)).first()

                    doctor_id = doctor_obj.id if doctor_obj else None

                    

                    # 8. Clinic

                    clinic_name = str(row[8]).strip() if len(row) > 8 and row[8] else None

                    clinic_id = None

                    if clinic_name:

                        clinic = Clinic.query.filter(Clinic.name.ilike(clinic_name)).first()

                        if clinic:

                            clinic_id = clinic.id

                            

                    # 9. Payment Method

                    payment_name = str(row[14]).strip() if len(row) > 14 and row[14] else None

                    payment_method_id = None

                    if payment_name:

                        pm = PaymentMethod.query.filter(PaymentMethod.name.ilike(payment_name)).first()

                        if pm:

                            payment_method_id = pm.id

                            

                    # 10. Contract

                    contract = str(row[3]).strip() if len(row) > 3 and row[3] else None

                    

                    # 11. Child

                    is_child_val = str(row[5]).strip().upper() if len(row) > 5 and row[5] else "FALSE"

                    is_child = (is_child_val == 'TRUE' or is_child_val == 'YES' or is_child_val == '1')

                    

                    # 12. Lab Tech

                    lab_tech_val = str(row[2]).strip() if len(row) > 2 and row[2] else None

                    

                    comment = str(row[17]).strip() if len(row) > 17 and row[17] else ""

                             

                    # Create Appointment

                    appt = Appointment(

                        patient_name=patient_name,

                        patient_phone='-', # No phone in excel

                        doctor=doctor_name,

                        doctor_id=doctor_id,

                        service=service_name,

                        date=appt_date,

                        time="00:00", # Default

                        author_id=current_user.id,

                        clinic_id=clinic_id,

                        center_id=int(center_id),

                        contract_number=contract,

                        quantity=qty,

                        cost=final_cost,

                        discount=discount,

                        comment=comment,

                        lab_tech=lab_tech_val,

                        payment_method_id=payment_method_id,

                        is_child=is_child

                    )

                    db.session.add(appt)

                    created_count += 1

                

                db.session.commit()

                flash(f'Загружено записей: {created_count}', 'success')

                

             else:

                 flash('Только .xlsx файлы поддерживаются для журнала', 'error')



        except Exception as e:

            db.session.rollback()

            flash(f'Ошибка импорта: {str(e)}', 'error')



    return redirect(url_for('main.journal', center_id=center_id))



# --- Clinic Management ---



@admin.route('/clinics')

def clinics():

    # Use Clinic model

    clinics = Clinic.query.order_by(Clinic.name.asc()).all()

    cities = Location.query.filter_by(type='city').all()

    return render_template('admin_clinics.html', clinics=clinics, cities=cities)



@admin.route('/clinics/add', methods=['POST'])

def add_clinic():

    name = request.form.get('name')

    city_id = request.form.get('city_id')

    phone = request.form.get('phone')

    

    if not name or not city_id:

        flash('Название и Город обязательны', 'error')

        return redirect(url_for('admin.clinics'))

        

    clinic = Clinic(

        name=name,

        city_id=int(city_id),

        phone=phone,

        is_cashless=(request.form.get('is_cashless') == 'on')

    )

    db.session.add(clinic)

    db.session.commit()

    flash(f'Клиника {name} добавлена', 'success')

    return redirect(url_for('admin.clinics'))



@admin.route('/clinics/<int:id>/update', methods=['POST'])

def update_clinic(id):

    clinic = Clinic.query.get_or_404(id)

    

    clinic.name = request.form.get('name')

    clinic.city_id = request.form.get('city_id')

    clinic.phone = request.form.get('phone')

    clinic.is_cashless = (request.form.get('is_cashless') == 'on')

    

    db.session.commit()

    flash(f'Клиника {clinic.name} обновлена', 'success')

    return redirect(url_for('admin.clinics'))



@admin.route('/clinics/<int:id>/delete', methods=['POST'])

def delete_clinic(id):

    clinic = Clinic.query.get_or_404(id)

    db.session.delete(clinic)

    db.session.commit()

    flash('Клиника удалена', 'success')

    return redirect(url_for('admin.clinics'))



@admin.route('/clinics/import', methods=['POST'])

def import_clinics():

    if 'file' not in request.files:

        flash('Нет файла', 'error')

        return redirect(url_for('admin.clinics'))

    

    file = request.files['file']

    if file.filename == '':

        flash('Файл не выбран', 'error')

        return redirect(url_for('admin.clinics'))



    if file:

        filename = secure_filename(file.filename)

        clinics_created = 0

        cities_created = 0

        

        try:

            # Determine file type

            if filename.endswith('.csv'):

                # Handle CSV

                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)

                csv_input = csv.reader(stream)

                header = next(csv_input, None)

                

                for row in csv_input:

                    if len(row) >= 1: # Name required

                        name = row[0].strip()

                        if not name: continue

                        city_name = row[1].strip() if len(row) > 1 else None

                        phone = row[2].strip() if len(row) > 2 else None

                        

                        # Process City

                        city = None

                        if city_name:

                             city = Location.query.filter(Location.name.ilike(city_name)).first()

                             if not city:

                                 city = Location(name=city_name)

                                 db.session.add(city)

                                 db.session.flush() # Get ID

                                 cities_created += 1



                        # Check exist

                        if not Clinic.query.filter_by(name=name).first():

                            # If city is None, we can't create Clinic because city_id is nullable=False in Model?

                            # Model says: city_id = db.Column(..., nullable=False)

                            # So we MUST have a city. 

                            # If no city in file, what to do? Skip or Default?

                            # Skip for now if no city.

                            if city:

                                clinic = Clinic(name=name, city_id=city.id, phone=phone)

                                db.session.add(clinic)

                                clinics_created += 1

                        

            elif filename.endswith('.xlsx'):

                # Handle Excel

                wb = openpyxl.load_workbook(file)

                sheet = wb.active

                

                # Iterate rows, skipping header (min_row=2)

                for row in sheet.iter_rows(min_row=2, values_only=True):

                    if not row or not row[0]: continue

                    name = str(row[0]).strip()

                    city_name = str(row[1]).strip() if len(row) > 1 and row[1] else None

                    phone = str(row[2]).strip() if len(row) > 2 and row[2] else None

                    

                    # Process City

                    city = None

                    if city_name:

                         city = Location.query.filter(Location.name.ilike(city_name)).first()

                         if not city:

                             city = Location(name=city_name)

                             db.session.add(city)

                             db.session.flush() # Get ID

                             cities_created += 1

                    

                    if not Clinic.query.filter_by(name=name).first():

                        if city: # City required

                            clinic = Clinic(name=name, city_id=city.id, phone=phone)

                            db.session.add(clinic)

                            clinics_created += 1



            else:

                 flash('Неподдерживаемый формат файла. Используйте CSV или XLSX.', 'error')

                 return redirect(url_for('admin.clinics'))



            db.session.commit()

            flash(f'Импортировано клиник: {clinics_created} (Новых городов: {cities_created})', 'success')



        except Exception as e:

            db.session.rollback()

            flash(f'Ошибка импорта: {str(e)}', 'error')

            

    return redirect(url_for('admin.clinics'))

# --- ICS Import Logic ---

from app.utils.ics_utils import parse_ics_content

import difflib



@admin.route('/import_ics', methods=['GET', 'POST'])

@login_required

def import_ics():

    centers = Location.query.filter_by(type='center').all()

    

    # Default start date: 2 months ago

    from datetime import timedelta

    default_start_date = (date.today() - timedelta(days=60)).strftime('%Y-%m-%d')

    

    if request.method == 'POST':

        if 'ics_file' not in request.files:

            flash('Нет файла', 'error')

            return redirect(request.url)

            

        file = request.files['ics_file']

        if file.filename == '':

            flash('Нет выбранного файла', 'error')

            return redirect(request.url)

            

        center_id = request.form.get('center_id')

        if not center_id:

             flash('Выберите филиал', 'error')

             return redirect(request.url)



        start_date_str = request.form.get('start_date')

        filter_date = None

        if start_date_str:

            try:

                filter_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()

            except ValueError:

                pass

        

        if file:

            try:

                content = file.read().decode('utf-8')

                parsed_events = parse_ics_content(content)

                

                # --- Filter by Date ---

                if filter_date:

                    filtered_events = []

                    for event in parsed_events:

                        evt_date = datetime.strptime(event['date'], '%Y-%m-%d').date()

                        if evt_date >= filter_date:

                            filtered_events.append(event)

                    parsed_events = filtered_events

                

                # --- Fuzzy Matching Context ---

                services = Service.query.all()

                doctors = Doctor.query.all()

                

                # Pre-calculate normalized names for matching

                def normalize(s): return s.lower().replace(' ', '') if s else ''

                

                svc_map = {normalize(s.name): s.id for s in services}

                doc_map = {normalize(d.name): d.id for d in doctors}

                

                # Prepare enriched events for preview

                for event in parsed_events:

                    summary = event['raw_summary']

                    norm_summary = normalize(summary)

                    

                    # 1. Match Service

                    event['matched_service_id'] = None

                    # Try direct fuzzy on keywords or specific known substrings

                    # Simple heuristic: Check if any service name is substring of summary

                    best_ratio = 0

                    best_svc_id = None

                    

                    for s in services:

                        # Check "OPTG", "KT", "ORT" special logic because user files use shorthand

                        # This is specific custom logic based on observation

                        s_norm = normalize(s.name)

                        if len(s_norm) < 3: continue # Skip short

                        

                        # Direct containment

                        if s_norm in norm_summary:

                            best_ratio = 1.0

                            best_svc_id = s.id

                            break

                            

                        # Fuzzy

                        ratio = difflib.SequenceMatcher(None, s_norm, norm_summary).ratio()

                        # If summary is long, ratio might be low even if match is good.

                        # Better metric: check word overlap?

                        pass 

                        

                    # Improved Heuristic for key types

                    if not best_svc_id:

                        summary_lower = summary.lower()

                        # Keywords mapping to DB names (approximate)

                        keyword_map = {

                            'оптг': 'ОПТГ', 

                            'кт': 'КТ', 

                            'орт': 'Ортодонт' # Just guessing ID or Name pattern

                        }

                        for kw, target_name in keyword_map.items():

                            if kw in summary_lower:

                                # Find service with this name

                                matched_svc = next((s for s in services if target_name.lower() in s.name.lower()), None)

                                if matched_svc:

                                    best_svc_id = matched_svc.id

                                    break

                    

                    event['matched_service_id'] = best_svc_id

                    

                    # 2. Match Doctor

                    # Use extracted candidate or scan summary

                    event['matched_doctor_id'] = None

                    doc_candidate = event.get('doctor_from_desc')

                    

                    if doc_candidate:

                        # Fuzzy match candidate against DB

                        matches = difflib.get_close_matches(doc_candidate, [d.name for d in doctors], n=1, cutoff=0.6)

                        if matches:

                            target_doc = next(d for d in doctors if d.name == matches[0])

                            event['matched_doctor_id'] = target_doc.id

                    

                    # Fallback: scan summary for doctor names? (Might be expensive/risky)

                

                return render_template(

                    'import_ics.html', 

                    parsed_events=parsed_events, 

                    centers=centers, 

                    services=services, 

                    doctors=doctors,

                    center_id=center_id

                )

                

            except Exception as e:

                flash(f'Ошибка обработки файла: {str(e)}', 'error')

                return redirect(request.url)



    return render_template('import_ics.html', centers=centers, default_start_date=default_start_date)



@admin.route('/import_ics/confirm', methods=['POST'])

@login_required

def confirm_ics_import():

    center_id = request.form.get('center_id')

    

    # Process form list data

    # format: events[0][date], events[0][skip]...

    # Flask doesn't parse nested dicts automatically well for arbitrary lists this way easily without third party lib or manual loop

    # We will manually iterate since we know the structure or expected index count?

    # Actually, easier to iterate keys.

    

    imported_count = 0

    

    # Reconstruct data from flat keys

    events_data = {}

    for key, value in request.form.items():

        if key.startswith('events['):

            # Key format: events[0][field_name]

            try:

                _, index_str, field_part = key.split('[')

                index = int(index_str[:-1]) # remove ]

                field = field_part[:-1] # remove ]

                

                if index not in events_data:

                    events_data[index] = {}

                events_data[index][field] = value

            except ValueError:

                continue

                

    for index, data in events_data.items():

        if 'skip' in data and data['skip'] == '1':

            continue

            

        try:

            date_obj = datetime.strptime(data['date'], '%Y-%m-%d').date()

            time_str = data['time']

            

            # Create Appointment

            new_appt = Appointment(

                center_id=int(center_id),

                date=date_obj,

                time=time_str,

                patient_name=data.get('patient_name', 'Unknown'),

                patient_phone=data.get('patient_phone', ''),

                doctor_id=int(data['doctor_id']) if data.get('doctor_id') else None,

                # Simple service handling (not linking M2M fully if just quick import, or link first one)

                # Ideally we link AppointmentService

                quantity=1,

                author_id=current_user.id

            )

            

            db.session.add(new_appt)

            db.session.flush() # to get ID

            

            svc_id = data.get('service_id')

            if svc_id:

                svc = Service.query.get(int(svc_id))

                if svc:

                    # Link

                    assoc = AppointmentService(service_id=svc.id, appointment_id=new_appt.id, quantity=1)

                    db.session.add(assoc)

                    # Also update summary string

                    new_appt.service = svc.name

            else:

                new_appt.service = "Импорт (Неизвестно)"

            

            imported_count += 1

            

        except Exception as e:

            # Continue or fail?

            # Let's log/print and continue for robust partial import

            print(f"Failed to import row {index}: {e}")

            continue



    try:

        db.session.commit()

        flash(f'Успешно импортировано записей: {imported_count}', 'success')

    except Exception as e:

        db.session.rollback()

        flash(f'Ошибка сохранения: {str(e)}', 'error')

        

    return redirect(url_for('admin.import_ics'))


@admin.route('/users/impersonate/<int:user_id>')
@login_required
def impersonate_user(user_id):
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
        
    user = User.query.get_or_404(user_id)
    login_user(user)
    flash(f'Вы вошли как {user.username}', 'success')
    return redirect(url_for('main.dashboard'))

@admin.route('/reports/api/organizations')
@login_required
def reports_organizations():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    month_str = request.args.get('month')
    if not month_str:
        today = datetime.now()
        month_str = today.strftime('%Y-%m')
        
    try:
        start_date = datetime.strptime(month_str, '%Y-%m').date()
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1, day=1)
    except ValueError:
        return jsonify({'error': 'Invalid month format'}), 400

    stats = db.session.query(
        User.id,
        User.username, 
        db.func.count(Appointment.id).label('count')
    ).join(Appointment, Appointment.author_id == User.id)\
     .filter(User.role == 'org')\
     .filter(Appointment.date >= start_date, Appointment.date < end_date)\
     .group_by(User.id)\
     .all()
     
    data = [{'id': s.id, 'username': s.username, 'count': s.count} for s in stats]
    return jsonify(data)

@admin.route('/reports/api/organizations/details')
@login_required
def reports_organizations_details():
    user_id = request.args.get('user_id', type=int)
    
    if current_user.role != 'superadmin':
        if current_user.role != 'org' or user_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        month_str = request.args.get('month')
        
        if not user_id or not month_str:
            return jsonify({'error': 'Missing parameters'}), 400
            
        try:
            start_date = datetime.strptime(month_str, '%Y-%m').date()
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1, day=1)
        except ValueError:
            return jsonify({'error': 'Invalid month format'}), 400

        # Fetch appointments
        appts = Appointment.query.filter_by(author_id=user_id)\
                .filter(Appointment.date >= start_date, Appointment.date < end_date)\
                .order_by(Appointment.date, Appointment.time).all()

        if not appts:
            return jsonify([])

        # To check "In Journal", we look for registered appointments on the same day for the same patient.
        relevant_dates = list({a.date for a in appts})
        
        registered_appts_on_dates = Appointment.query.filter(
            Appointment.date.in_(relevant_dates),
            Appointment.payment_method_id != None
        ).all()

        # Helper for name normalization
        import difflib
        def normalize_name(name):
            return name.lower().replace(' ', '') if name else ''

        results = []
        for i, appt in enumerate(appts, 1):
            is_registered = False
            
            # 1. Direct check
            if appt.payment_method_id is not None:
                is_registered = True
            else:
                # 2. Fuzzy match with registered appointments on the same day
                appt_norm_name = normalize_name(appt.patient_name)
                if appt_norm_name:
                    for reg in registered_appts_on_dates:
                        if reg.date == appt.date:
                            reg_norm_name = normalize_name(reg.patient_name)
                            if appt_norm_name == reg_norm_name:
                                is_registered = True
                                break
                            if difflib.SequenceMatcher(None, appt_norm_name, reg_norm_name).ratio() > 0.85:
                                is_registered = True
                                break
            
            results.append({
                'n_pp': i,
                'patient_name': appt.patient_name,
                'center_name': appt.center.name if appt.center else 'Unknown',
                'date': appt.date.strftime('%d.%m.%Y'),
                'time': appt.time,
                'is_registered': is_registered
            })

        return jsonify(results)
    except Exception as e:
        print(f"Error in reports_organizations_details: {e}")
        return jsonify({'error': str(e)}), 500

@admin.route('/reports/api/lab_techs')
@login_required
def reports_lab_techs():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    # Lab Techs are users with role 'lab_tech' OR 'superadmin' (except specific 'admin' user)
    # Workload: Patient Count, Revenue
    
    year_str = request.args.get('year')
    month_str = request.args.get('month')
    day_str = request.args.get('day')
    
    # Default to current month if nothing provided
    if not year_str and not month_str and not day_str:
        today = datetime.now()
        year_str = str(today.year)
        month_str = str(today.month).zfill(2)
        
    query = db.session.query(
        User.username,
        db.func.count(Appointment.id).label('appt_count'),
        db.func.sum(Appointment.cost).label('total_revenue'),
        db.func.count(db.func.distinct(Appointment.patient_name)).label('unique_patients')
    ).join(Appointment, Appointment.author_id == User.id)\
     .filter(User.role.in_(['lab_tech', 'superadmin']))\
     .filter(User.username != 'admin')

    # Date Filtering
    try:
        if year_str and month_str and day_str and day_str != '00':
            # Specific Day
            target_date = datetime.strptime(f"{year_str}-{month_str}-{day_str}", '%Y-%m-%d').date()
            query = query.filter(db.func.date(Appointment.date) == target_date)
        elif year_str and month_str:
            # Whole Month
            start_date = datetime.strptime(f"{year_str}-{month_str}", '%Y-%m').date()
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1, day=1)
            query = query.filter(Appointment.date >= start_date, Appointment.date < end_date)
        elif year_str:
            # Whole Year
            start_date = datetime(int(year_str), 1, 1).date()
            end_date = datetime(int(year_str) + 1, 1, 1).date()
            query = query.filter(Appointment.date >= start_date, Appointment.date < end_date)
    except ValueError:
        pass # Ignore invalid dates and define no filter (or return error?)

    stats = query.group_by(User.id).all()
     
    data = [{
        'username': s.username, 
        'appt_count': s.appt_count,
        'total_revenue': s.total_revenue or 0,
        'unique_patients': s.unique_patients
    } for s in stats]
    
    return jsonify(data)

@admin.route('/reports/api/comparative')
@login_required
def reports_comparative():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403

    year_str = request.args.get('year')
    month_str = request.args.get('month')
    day_str = request.args.get('day')

    if not year_str:
        return jsonify({'error': 'Year is required'}), 400

    def get_period_stats(year, month, day):
        # date filtering logic
        start = None
        end = None
        try:
            if month and day and day != '00':
                start = datetime.strptime(f"{year}-{month}-{day}", '%Y-%m-%d').date()
                end = start + timedelta(days=1)
            elif month and month != '00':
                start = datetime.strptime(f"{year}-{month}", '%Y-%m').date()
                if start.month == 12:
                    end = start.replace(year=start.year + 1, month=1, day=1)
                else:
                    end = start.replace(month=start.month + 1, day=1)
            else:
                start = datetime(int(year), 1, 1).date()
                end = datetime(int(year) + 1, 1, 1).date()
        except Exception as e:
            print(f"Error in get_period_stats: {e}")
            return None

        # Query stats per center
        stats = db.session.query(
            Location.id.label('center_id'),
            Location.name.label('center_name'),
            db.func.count(db.func.distinct(
                db.case(
                    (Appointment.payment_method_id != None, Appointment.patient_name),
                    else_=None
                )
            )).label('patient_count'),
            db.func.sum(
                db.case(
                    (db.func.lower(PaymentMethod.name).in_(['наличные', 'карта']), Appointment.cost),
                    else_=0
                )
            ).label('total_sum')
        ).select_from(Location)\
         .outerjoin(Appointment, (Appointment.center_id == Location.id) & (Appointment.date >= start) & (Appointment.date < end))\
         .outerjoin(PaymentMethod, Appointment.payment_method_id == PaymentMethod.id)\
         .filter(Location.type == 'center')\
         .group_by(Location.id, Location.name).all()

        # Lab techs for chosen date (only if day is selected)
        labs_per_center = {}
        if day and day != '00':
            # 1. Names from the string field 'lab_tech'
            lab_field_query = db.session.query(
                Appointment.center_id,
                Appointment.lab_tech
            ).filter(Appointment.date == start)\
             .filter(Appointment.payment_method_id != None)\
             .filter(Appointment.lab_tech != None)\
             .filter(Appointment.lab_tech != '').distinct().all()
            
            # 2. Usernames from ALL authors of registered appointments
            lab_auth_query = db.session.query(
                Appointment.center_id,
                User.username
            ).join(User, Appointment.author_id == User.id)\
             .filter(Appointment.date == start)\
             .filter(Appointment.payment_method_id != None).distinct().all()

            for cid, val in lab_field_query:
                if cid not in labs_per_center:
                    labs_per_center[cid] = set()
                labs_per_center[cid].add(val)
            
            for cid, val in lab_auth_query:
                if cid not in labs_per_center:
                    labs_per_center[cid] = set()
                labs_per_center[cid].add(val)

        res_data = {}
        for s in stats:
            res_data[s.center_id] = {
                'name': s.center_name,
                'patient_count': s.patient_count,
                'total_sum': float(s.total_sum or 0),
                'labs': list(labs_per_center.get(s.center_id, []))
            }
        return res_data

    current_data = get_period_stats(year_str, month_str, day_str)
    last_year_data = get_period_stats(str(int(year_str) - 1), month_str, day_str)

    if current_data is None:
        return jsonify({'error': 'Invalid current date parameters'}), 400
    
    if last_year_data is None:
        last_year_data = {}

    # Combine
    centers_result = []
    total_patients_curr = 0
    total_sum_curr = 0
    total_patients_prev = 0
    total_sum_prev = 0

    try:
        for cid in current_data:
            curr = current_data[cid]
            prev = last_year_data.get(cid, {'patient_count': 0, 'total_sum': 0, 'labs': []})
            
            centers_result.append({
                'id': cid,
                'name': curr['name'],
                'labs': curr['labs'],
                'current_patients': curr['patient_count'],
                'current_sum': curr['total_sum'],
                'prev_patients': prev.get('patient_count', 0),
                'prev_sum': prev.get('total_sum', 0)
            })

            total_patients_curr += (curr['patient_count'] or 0)
            total_sum_curr += (curr['total_sum'] or 0)
            total_patients_prev += (prev.get('patient_count', 0) or 0)
            total_sum_prev += (prev.get('total_sum', 0) or 0)
    except Exception as e:
        print(f"Error combining comparative data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'centers': centers_result,
        'totals': {
            'current_patients': total_patients_curr,
            'current_money': total_sum_curr,
            'prev_patients': total_patients_prev,
            'prev_money': total_sum_prev
        }
    })

@admin.route('/reports/api/audit')
@login_required
def reports_audit():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    # Recent 100 actions
    history = AppointmentHistory.query.order_by(AppointmentHistory.timestamp.desc()).limit(100).all()
    
    data = []
    for h in history:
        data.append({
            'timestamp': h.timestamp.isoformat(),
            'user': h.user.username if h.user else 'Unknown',
            'action': h.action,
            'details': f"Appt #{h.appointment_id}"
        })
        
    return jsonify(data)

@admin.route('/reports/api/bonuses')
@login_required
def reports_bonuses():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    month_str = request.args.get('month')
    filter_type = request.args.get('filter_type', 'all') # all, with, without
    
    if not month_str:
        today = datetime.now()
        month_str = today.strftime('%Y-%m')
        
    try:
        start_date = datetime.strptime(month_str, '%Y-%m').date()
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1, day=1)
    except ValueError:
        return jsonify({'error': 'Invalid month format'}), 400
        
    # 1. Get Bonus Config
    period = BonusPeriod.query.filter(
        BonusPeriod.start_date <= start_date
    ).filter(
        (BonusPeriod.end_date >= end_date) | (BonusPeriod.end_date == None)
    ).order_by(BonusPeriod.start_date.desc()).first()
    
    bonus_map = {}
    if period:
        for bv in period.values:
            bonus_map[f"{bv.service_id}_{bv.column_index}"] = bv.value

    doctor_name_expr = db.func.coalesce(Doctor.name, Appointment.doctor, 'Unknown')
    
    try:
        # 2. Fetch all services performed in range
        query = db.session.query(
            Appointment.id,
            doctor_name_expr.label('doctor_name'),
            Doctor.bonus_type,
            Service.id.label('service_id'),
            AppointmentService.quantity
        ).select_from(Appointment)\
         .outerjoin(Doctor, Appointment.doctor_id == Doctor.id)\
         .outerjoin(AppointmentService, Appointment.id == AppointmentService.appointment_id)\
         .outerjoin(Service, AppointmentService.service_id == Service.id)\
         .filter(Appointment.date >= start_date, Appointment.date < end_date)
         
        # Apply Filter
        if filter_type == 'with':
            query = query.filter(Doctor.bonus_type != None)
        elif filter_type == 'without':
            query = query.filter(Doctor.bonus_type == None)
            
        results = query.all()
         
        # Aggregation
        stats = {} # name -> { appt_ids: set(), total_bonus: 0.0 }
        
        for r in results:
            name = r.doctor_name
            if not name or not name.strip(): name = 'Не указан'
            
            if name not in stats:
                stats[name] = {'appt_ids': set(), 'total_bonus': 0.0}
            
            stats[name]['appt_ids'].add(r.id)
            
            # Calc Bonus
            if r.bonus_type and period and r.service_id:
                key = f"{r.service_id}_{r.bonus_type}"
                val = bonus_map.get(key, 0.0)
                stats[name]['total_bonus'] += (val * (r.quantity or 1))
                
        items = []
        for name, data in stats.items():
            items.append({
                'name': name,
                'count': len(data['appt_ids']),
                'revenue': data['total_bonus'] # Using 'revenue' field for Total Bonus to minimize frontend change if logical
            })
            
        # Sort by count desc
        items.sort(key=lambda x: x['count'], reverse=True)
            
        return jsonify({
            'rows': items,
            'month': month_str
        })
    except Exception as e:
        print(f"Error in reports_bonuses: {e}")
        return jsonify({'error': str(e)}), 500

@admin.route('/reports/api/bonuses/details')
@login_required
def reports_bonuses_details():
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    month_str = request.args.get('month')
    doctor_name = request.args.get('doctor_name')
    
    if not month_str or not doctor_name:
        return jsonify({'error': 'Missing parameters'}), 400
        
    try:
        start_date = datetime.strptime(month_str, '%Y-%m').date()
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1, day=1)
    except ValueError:
        return jsonify({'error': 'Invalid month format'}), 400
        
    doctor_name_expr = db.func.coalesce(Doctor.name, Appointment.doctor, 'Unknown')
    
    try:
        # Fetch relevant Period for this month (using start_date)
        period = BonusPeriod.query.filter(
            BonusPeriod.start_date <= start_date
        ).filter(
            (BonusPeriod.end_date >= end_date) | (BonusPeriod.end_date == None)
        ).order_by(BonusPeriod.start_date.desc()).first()
        
        # Pre-fetch bonus values map: { service_id_col_index: value }
        bonus_map = {}
        if period:
            for bv in period.values:
                bonus_map[f"{bv.service_id}_{bv.column_index}"] = bv.value

        # Query: Select Patient Name, Service Name, Service ID, Doctor Bonus Type
        results = db.session.query(
            Appointment.patient_name,
            Appointment.date,
            Service.name.label('service_name'),
            Service.id.label('service_id'),
            Doctor.bonus_type
        ).select_from(Appointment)\
         .outerjoin(Doctor, Appointment.doctor_id == Doctor.id)\
         .outerjoin(AppointmentService, Appointment.id == AppointmentService.appointment_id)\
         .outerjoin(Service, AppointmentService.service_id == Service.id)\
         .filter(Appointment.date >= start_date, Appointment.date < end_date)\
         .filter(doctor_name_expr == doctor_name)\
         .order_by(Appointment.date.desc())\
         .all()
         
        details = []
        for r in results:
            bonus_amount = 0.0
            if r.bonus_type and period and r.service_id:
                key = f"{r.service_id}_{r.bonus_type}"
                bonus_amount = bonus_map.get(key, 0.0)
                
            details.append({
                'patient_name': r.patient_name,
                'date': r.date.strftime('%d.%m.%Y'),
                'service_name': r.service_name or '-',
                'bonus': bonus_amount
            })
            
        return jsonify(details)
    except Exception as e:
        print(f"Error in reports_bonuses_details: {e}")
        return jsonify({'error': str(e)}), 500

@admin.route('/reports/today')
@admin.route('/reports')
@login_required
def reports_today():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('reports_today.html')

@admin.route('/reports/organizations')
@login_required
def reports_organizations_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('reports_org_activity.html')

@admin.route('/reports/lab_workload')
@login_required
def reports_lab_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('reports_lab_workload.html')

@admin.route('/reports/logs')
@login_required
def reports_logs_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('reports_logs.html')

@admin.route('/reports/bonuses')
@login_required
def reports_bonuses_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    
    return render_template('reports_bonuses.html')

@admin.route('/reports/cashless')
@login_required
def reports_cashless_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    
    return render_template('reports_cashless.html')

@admin.route('/reports/bonuses/config')
@login_required
def reports_bonuses_config_page():
    if current_user.role != 'superadmin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('main.dashboard'))
    
    services = Service.query.order_by(Service.name).all()
    services_data = [{'id': s.id, 'name': s.name} for s in services]
    return render_template('reports_bonuses_config.html', services=services_data)

@admin.route('/api/bonuses/config', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def api_bonuses_config():
    from app.models import BonusPeriod, BonusValue # Import here to avoid circular
    from datetime import datetime
    
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Unauthorized'}), 403

    if request.method == 'GET':
        periods = BonusPeriod.query.order_by(BonusPeriod.start_date).all()
        return jsonify([p.to_dict() for p in periods])

    if request.method == 'POST':
        data = request.json
        try:
            # Full sync: delete all and recreate
            # Delete children first to avoid FK constraint violations
            BonusValue.query.delete()
            BonusPeriod.query.delete()
            
            for p_data in data:
                period = BonusPeriod(
                    start_date=datetime.strptime(p_data['startDate'], '%Y-%m-%d').date(),
                    end_date=datetime.strptime(p_data['endDate'], '%Y-%m-%d').date() if p_data.get('endDate') else None,
                    columns=int(p_data.get('columns', 1)) 
                )
                db.session.add(period)
                db.session.flush() # Generate ID
                
                if 'values' in p_data:
                    for v_data in p_data['values']:
                        try:
                            val = float(v_data.get('val', 0))
                        except:
                            val = 0
                        
                        if val != 0: 
                            b_val = BonusValue(
                                period_id=period.id,
                                service_id=v_data['serviceId'],
                                column_index=v_data['col'],
                                value=val
                            )
                            db.session.add(b_val)
            
            db.session.commit()
            return jsonify({'status': 'success'})
        except Exception as e:
            db.session.rollback()
            print(f"Error saving bonuses config: {e}") # Log to console
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

@admin.route('/notifications', methods=['GET', 'POST'])
@login_required
def notifications():
    if request.method == 'POST':
        title = request.form.get('title')
        message = request.form.get('message')
        
        target_type = request.form.get('target_type')  # 'all', 'role', 'user'
        target_role = request.form.get('target_role')
        target_user_id = request.form.get('target_user_id')

        if not title or not message:
            flash('Заполните все поля', 'error')
            return redirect(url_for('admin.notifications'))

        # Create Notification
        notif = Notification(
            title=title,
            message=message,
            target_type=target_type,
            target_value=target_role if target_type == 'role' else (target_user_id if target_type == 'user' else None),
            author_id=current_user.id
        )
        db.session.add(notif)
        db.session.flush()

        # Determine Recipients
        recipients = []
        if target_type == 'all':
            recipients = User.query.filter_by(is_blocked=False).all()
        elif target_type == 'role':
            if target_role:
                recipients = User.query.filter_by(role=target_role, is_blocked=False).all()
        elif target_type == 'user':
            if target_user_id:
                recipients = User.query.filter_by(id=target_user_id).all()

        # Create Statuses
        count = 0
        for user in recipients:
            status = NotificationStatus(
                notification_id=notif.id,
                user_id=user.id
            )
            db.session.add(status)
            count += 1
        
        try:
            db.session.commit()
            flash(f'Уведомление отправлено {count} пользователям!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка отправки: {str(e)}', 'error')

        return redirect(url_for('admin.notifications'))
    
    # GET: Prepare data
    users = User.query.filter_by(is_blocked=False).order_by(User.username).all()
    roles = ['superadmin', 'admin', 'org', 'lab_tech', 'doctor'] # Fixed list or derived? Existing roles.
    
    # History (Last 20)
    history = Notification.query.order_by(Notification.created_at.desc()).limit(20).all()
    
    # Enrich history with stats
    history_data = []
    for n in history:
        total = NotificationStatus.query.filter_by(notification_id=n.id).count()
        read = NotificationStatus.query.filter_by(notification_id=n.id, is_read=True).count()
        history_data.append({
            'notif': n,
            'total': total,
            'read': read,
            'percent': int((read/total)*100) if total > 0 else 0
        })

    return render_template('admin_notifications.html', users=users, roles=roles, history=history_data)
